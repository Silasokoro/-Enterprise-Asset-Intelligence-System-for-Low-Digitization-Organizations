asset_disposal_model.py
import time
import pandas as pd
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook

# ============================================
# FILE CONFIGURATION
# ============================================

FILE_PATH = r"Asset_Intelligence_System.xlsx"
SHEET_NAME = "Asset_Master_List"

# ============================================
# DISPOSAL MODEL FUNCTION
# ============================================

def run_disposal_model():

    print("\nReading updated asset master list...")

    # Load Excel sheet
    df = pd.read_excel(FILE_PATH, sheet_name=SHEET_NAME)

    # Clean numeric columns
    df['Acquisition_Cost'] = pd.to_numeric(
        df['Acquisition_Cost'],
        errors='coerce'
    )

    df['Total_Lifetime_Maintenance_Cost'] = pd.to_numeric(
        df['Total_Lifetime_Maintenance_Cost'],
        errors='coerce'
    )

    # Avoid division by zero
    df['Maintenance_Ratio'] = (
        df['Total_Lifetime_Maintenance_Cost'] /
        df['Acquisition_Cost']
    ).fillna(0)

    # ============================================
    # DISPOSAL LOGIC
    # ============================================

    def disposal_flag(ratio):

        if ratio >= 0.70:
            return 'DISPOSAL RECOMMENDED'

        elif ratio >= 0.50:
            return 'MONITOR CLOSELY'

        else:
            return 'CONTINUE USE'

    df['Disposal_Recommendation'] = df['Maintenance_Ratio'].apply(disposal_flag)

    # ============================================
    # OPTIONAL FINANCIAL RISK SCORE
    # ============================================

    df['Financial_Risk_Score'] = (
        df['Maintenance_Ratio'] * 100
    ).round(2)

    # ============================================
    # EXPORT BACK TO EXCEL
    # ============================================

    workbook = load_workbook(FILE_PATH)

    if 'Disposal_Model_Output' in workbook.sheetnames:
        del workbook['Disposal_Model_Output']

    workbook.save(FILE_PATH)

    with pd.ExcelWriter(
        FILE_PATH,
        engine='openpyxl',
        mode='a'
    ) as writer:

        df.to_excel(
            writer,
            sheet_name='Disposal_Model_Output',
            index=False
        )

    print("Disposal model updated successfully.")

    # ============================================
    # SUMMARY OUTPUT
    # ============================================

    summary = df['Disposal_Recommendation'].value_counts()

    print("\nASSET DISPOSAL SUMMARY")
    print(summary)

# ============================================
# FILE MONITORING AUTOMATION
# ============================================

class ExcelFileHandler(FileSystemEventHandler):

    def on_modified(self, event):

        if FILE_PATH in event.src_path:

            print("\nDetected update in asset master list...")

            try:
                run_disposal_model()

            except Exception as e:
                print(f"Error: {e}")

# ============================================
# MAIN PROGRAM
# ============================================

if __name__ == '__main__':

    # Run immediately at startup
    run_disposal_model()

    # Watch file for changes
    event_handler = ExcelFileHandler()

    observer = Observer()

    observer.schedule(event_handler, path='.', recursive=False)

    observer.start()

    print("\nMonitoring Excel file for updates...")
    print("Press CTRL+C to stop monitoring.")

    try:
        while True:
            time.sleep(5)

    except KeyboardInterrupt:
        observer.stop()

    observer.join()
________________________________________
